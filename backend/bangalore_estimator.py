"""Bangalore-specific medical cost estimator backed by ``blr.xlsx``.

Why this module exists
======================

The generic Karnataka estimator (:mod:`cost_estimator`) is deterministic and
anchored on a hand-calibrated ``condition × tier × severity`` price table plus
a small ``karnataka_hospitals_200.csv`` directory. For **Bangalore /
Bengaluru** we have a far richer dataset — ``blr.xlsx`` — with 300 hospitals,
~5,000 individual doctors, real per-doctor *consultation fees*,
specializations, qualifications, experience, accreditation, bed counts, and
availability.

This module turns that dataset into a localized, fee-driven estimation +
recommendation pipeline that is intentionally *richer* than the generic one:

  * **Dynamic tier classification.** Each hospital is bucketed into
    ``Low / Mid / High`` pricing tiers from the *distribution* of its
    consultation fees (terciles computed at import time), not hardcoded
    bands.

  * **Doctor-specialization filtering.** A free-text condition is mapped to
    the relevant ``blr.xlsx`` specialty departments, and the most relevant
    doctors per hospital are surfaced with their fee, qualification, and
    experience.

  * **Fee-anchored cost ranges.** The consultation line is anchored to the
    *real* consultation fees observed in the dataset for the chosen tier,
    while the remaining structural lines (tests, medication, procedure,
    hospitalization) reuse the proven calibration in
    :mod:`cost_estimator`. This keeps the expensive surgical/hospital lines
    realistic while letting genuine Bangalore fee data drive the
    consultation cost and tier differentiation.

Everything is computed once at import time; the public functions are pure.
If ``blr.xlsx`` or ``openpyxl`` is unavailable the module degrades gracefully
(``AVAILABLE = False``) and callers fall back to the generic estimator.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cost_estimator
from cost_estimator import ConditionProfile

# New production-grade symptom -> specialization mapper. Imported defensively
# so a problem in the mapper never takes down the estimator; if unavailable we
# fall back to the legacy condition-key mapping below.
try:
    import specialization_mapper as _spec_mapper
    _MAPPER_AVAILABLE = True
except Exception:  # pragma: no cover - defensive
    _spec_mapper = None  # type: ignore[assignment]
    _MAPPER_AVAILABLE = False

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Dataset location
# ---------------------------------------------------------------------------
#
# ``blr.xlsx`` lives at the repository root alongside
# ``karnataka_hospitals_200.csv``. We probe a few sensible locations so the
# loader works both in local dev (repo root) and if the file is copied next
# to the backend for containerized deploys.

_DATASET_FILENAME = "blr.xlsx"
_DATASET_CANDIDATES: Tuple[Path, ...] = (
    Path(__file__).resolve().parent.parent / _DATASET_FILENAME,  # repo root
    Path(__file__).resolve().parent / _DATASET_FILENAME,         # backend/
)
_SHEET_NAME = "Hospital Directory"


def _resolve_dataset_path() -> Optional[Path]:
    for candidate in _DATASET_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


# ---------------------------------------------------------------------------
# City detection
# ---------------------------------------------------------------------------

_BANGALORE_ALIASES = frozenset(
    {
        "bengaluru",
        "bangalore",
        "bengaluru urban",
        "bengaluru rural",
        "blr",
        "bglr",
    }
)


def is_bangalore(city: Optional[str]) -> bool:
    """Return ``True`` when ``city`` refers to Bengaluru/Bangalore."""
    if not city:
        return False
    return city.strip().lower() in _BANGALORE_ALIASES


# ---------------------------------------------------------------------------
# Domain records
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Doctor:
    name: str
    specialty: str          # raw blr.xlsx specialty (e.g. "Cardiology")
    qualification: str
    experience_years: int
    consultation_fee: int
    availability: str
    timing: str


@dataclass
class BangaloreHospital:
    hospital_id: int
    name: str
    area: str
    address: str
    hospital_type: str       # Private / Corporate / Government / Trust / Charitable
    total_beds: int
    icu_beds: int
    established_year: int
    rating: float
    accreditation: str
    emergency: str
    ambulance: str
    specialties: Tuple[str, ...]
    doctors: Tuple[Doctor, ...]
    # Derived (populated after load):
    median_fee: int = 0
    min_fee: int = 0
    max_fee: int = 0
    tier: str = "Mid"        # Low | Mid | High


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _to_int(value: object, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        try:
            return int(round(float(value)))
        except (ValueError, OverflowError):
            return default
    digits = re.sub(r"[^0-9]", "", str(value))
    return int(digits) if digits else default


def _to_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(m.group(0)) if m else default


def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------


# Column headers expected in the "Hospital Directory" sheet. Resolved to
# positional indices at load time so minor column reordering won't break us.
_COLUMNS = (
    "Hospital ID",
    "Hospital Name",
    "Area",
    "Full Address",
    "Hospital Type",
    "Total Beds",
    "ICU Beds",
    "Operation Theatres",
    "Established Year",
    "Rating (5)",
    "Accreditation",
    "Emergency (24/7)",
    "Ambulance Service",
    "Phone",
    "Email",
    "Specialty/Department",
    "Doctor Name",
    "Qualification",
    "Experience",
    "Consultation Fee",
    "Availability",
    "Timing",
)


def _load_hospitals() -> Tuple[BangaloreHospital, ...]:
    path = _resolve_dataset_path()
    if path is None:
        logger.warning(
            "bangalore_estimator: %s not found in %s; Bangalore pipeline disabled",
            _DATASET_FILENAME,
            [str(p) for p in _DATASET_CANDIDATES],
        )
        return ()

    try:
        import openpyxl  # imported lazily so a missing dep doesn't crash import
    except Exception:  # pragma: no cover - environment dependent
        logger.warning(
            "bangalore_estimator: openpyxl unavailable; Bangalore pipeline disabled"
        )
        return ()

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        logger.exception("bangalore_estimator: failed to open %s", path)
        return ()

    try:
        if _SHEET_NAME in wb.sheetnames:
            ws = wb[_SHEET_NAME]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return ()

        # Map header label -> column index (case/space tolerant).
        norm_header = {
            _clean(h).lower(): idx for idx, h in enumerate(header) if h is not None
        }

        def col(label: str) -> Optional[int]:
            return norm_header.get(label.lower())

        idx = {name: col(name) for name in _COLUMNS}

        def cell(row: Tuple, name: str):
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        # Accumulate doctor rows grouped by hospital id (fallback: name).
        grouped: Dict[object, Dict[str, object]] = {}
        for row in rows:
            if row is None:
                continue
            name = _clean(cell(row, "Hospital Name"))
            if not name:
                continue
            hid = cell(row, "Hospital ID")
            key = hid if hid is not None else name

            bucket = grouped.get(key)
            if bucket is None:
                bucket = {
                    "hospital_id": _to_int(hid),
                    "name": name,
                    "area": _clean(cell(row, "Area")),
                    "address": _clean(cell(row, "Full Address")),
                    "hospital_type": _clean(cell(row, "Hospital Type")),
                    "total_beds": _to_int(cell(row, "Total Beds")),
                    "icu_beds": _to_int(cell(row, "ICU Beds")),
                    "established_year": _to_int(cell(row, "Established Year")),
                    "rating": _to_float(cell(row, "Rating (5)")),
                    "accreditation": _clean(cell(row, "Accreditation")),
                    "emergency": _clean(cell(row, "Emergency (24/7)")),
                    "ambulance": _clean(cell(row, "Ambulance Service")),
                    "doctors": [],
                    "specialties": set(),
                }
                grouped[key] = bucket

            specialty = _clean(cell(row, "Specialty/Department"))
            doctor_name = _clean(cell(row, "Doctor Name"))
            if doctor_name:
                doc = Doctor(
                    name=doctor_name,
                    specialty=specialty,
                    qualification=_clean(cell(row, "Qualification")),
                    experience_years=_to_int(cell(row, "Experience")),
                    consultation_fee=_to_int(cell(row, "Consultation Fee")),
                    availability=_clean(cell(row, "Availability")),
                    timing=_clean(cell(row, "Timing")),
                )
                bucket["doctors"].append(doc)  # type: ignore[union-attr]
            if specialty:
                bucket["specialties"].add(specialty)  # type: ignore[union-attr]
    finally:
        try:
            wb.close()
        except Exception:
            pass

    hospitals: List[BangaloreHospital] = []
    for bucket in grouped.values():
        doctors: List[Doctor] = list(bucket["doctors"])  # type: ignore[arg-type]
        fees = [d.consultation_fee for d in doctors if d.consultation_fee > 0]
        median_fee = int(_percentile(fees, 50)) if fees else 0
        hospitals.append(
            BangaloreHospital(
                hospital_id=int(bucket["hospital_id"]),  # type: ignore[arg-type]
                name=str(bucket["name"]),
                area=str(bucket["area"]),
                address=str(bucket["address"]),
                hospital_type=str(bucket["hospital_type"]),
                total_beds=int(bucket["total_beds"]),  # type: ignore[arg-type]
                icu_beds=int(bucket["icu_beds"]),  # type: ignore[arg-type]
                established_year=int(bucket["established_year"]),  # type: ignore[arg-type]
                rating=float(bucket["rating"]),  # type: ignore[arg-type]
                accreditation=str(bucket["accreditation"]),
                emergency=str(bucket["emergency"]),
                ambulance=str(bucket["ambulance"]),
                specialties=tuple(sorted(bucket["specialties"])),  # type: ignore[arg-type]
                doctors=tuple(doctors),
                median_fee=median_fee,
                min_fee=min(fees) if fees else 0,
                max_fee=max(fees) if fees else 0,
            )
        )
    return tuple(hospitals)


# ---------------------------------------------------------------------------
# Statistics helpers (no numpy dependency at runtime)
# ---------------------------------------------------------------------------


def _percentile(values: List[int], pct: float) -> float:
    """Linear-interpolation percentile (pct in 0..100). Empty -> 0.0."""
    return _percentile_f([float(v) for v in values], pct)


def _percentile_f(values: List[float], pct: float) -> float:
    """Linear-interpolation percentile over floats (pct in 0..100)."""
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return float(ordered[0])
    rank = (pct / 100.0) * (len(ordered) - 1)
    lo = int(rank)
    hi = min(lo + 1, len(ordered) - 1)
    frac = rank - lo
    return ordered[lo] * (1 - frac) + ordered[hi] * frac


# ---------------------------------------------------------------------------
# Module-level load + derived distributions
# ---------------------------------------------------------------------------

HOSPITALS: Tuple[BangaloreHospital, ...] = _load_hospitals()
AVAILABLE: bool = len(HOSPITALS) > 0

# Dynamic tier classification.
#
# Tiers are derived from a *pricing score* that blends two real signals from
# ``blr.xlsx``, exactly as the requirements call for ("consultation fee and
# hospital category/type as primary indicators"):
#
#   * the hospital's representative (median) consultation fee — weighted 70%,
#     ranked as a percentile across all hospitals so it is scale-free; and
#   * a prior implied by the hospital category/type — weighted 30% (public /
#     charitable institutions skew lower, corporate / private skew higher).
#
# The Low / Mid / High cut points are the terciles of the resulting score
# distribution, so nothing is hardcoded — re-running on a different dataset
# re-derives the boundaries automatically.
TIER_NAMES: Tuple[str, ...] = ("Low", "Mid", "High")

# Relative pricing prior per hospital category. Values are ordinal (0..1) and
# only matter relative to one another; they nudge the fee-based score so the
# tiers track real-world expectations when fees alone are ambiguous.
_TYPE_PRIOR: Dict[str, float] = {
    "government": 0.10,
    "charitable": 0.25,
    "trust": 0.40,
    "private": 0.65,
    "corporate": 0.85,
}
_FEE_WEIGHT = 0.70
_TYPE_WEIGHT = 0.30

_LOW_SCORE_CUT: float = 1.0 / 3.0
_HIGH_SCORE_CUT: float = 2.0 / 3.0
# Per-tier consultation fee anchor bands (25th..75th percentile of the
# per-hospital median fees within that tier). Because hospitals are tiered by
# a fee-weighted score, these bands come out cleanly separated Low < Mid <
# High and drive the fee-anchored consultation cost line.
TIER_FEE_BANDS: Dict[str, Tuple[int, int]] = {}


def _type_prior(hospital_type: str) -> float:
    return _TYPE_PRIOR.get((hospital_type or "").strip().lower(), 0.5)


def _build_tier_model() -> None:
    """Compute the pricing-score model, classify hospitals, and derive the
    per-tier consultation fee anchor bands — all from the dataset."""
    global _LOW_SCORE_CUT, _HIGH_SCORE_CUT, TIER_FEE_BANDS

    if not HOSPITALS:
        return

    medians = sorted(h.median_fee for h in HOSPITALS if h.median_fee > 0)
    if not medians:
        return

    def fee_rank(fee: int) -> float:
        """Percentile rank (0..1) of ``fee`` within the median-fee distribution."""
        if fee <= 0 or len(medians) <= 1:
            return 0.5
        below = sum(1 for m in medians if m < fee)
        return below / (len(medians) - 1)

    # Composite pricing score per hospital: fee rank (primary) + type prior.
    scores: List[float] = []
    score_by_id: Dict[int, float] = {}
    for h in HOSPITALS:
        score = _FEE_WEIGHT * fee_rank(h.median_fee) + _TYPE_WEIGHT * _type_prior(h.hospital_type)
        score_by_id[id(h)] = score
        scores.append(score)

    _LOW_SCORE_CUT = _percentile_f(scores, 100.0 / 3.0)
    _HIGH_SCORE_CUT = _percentile_f(scores, 200.0 / 3.0)

    for h in HOSPITALS:
        s = score_by_id[id(h)]
        if s <= _LOW_SCORE_CUT:
            h.tier = "Low"
        elif s >= _HIGH_SCORE_CUT:
            h.tier = "High"
        else:
            h.tier = "Mid"

    # Anchor bands from the per-hospital median fees inside each tier.
    for tier in TIER_NAMES:
        tier_medians = [h.median_fee for h in HOSPITALS if h.tier == tier and h.median_fee > 0]
        if tier_medians:
            lo = int(_percentile(tier_medians, 25))
            hi = int(_percentile(tier_medians, 75))
            TIER_FEE_BANDS[tier] = (lo, max(hi, lo + 1))

    all_medians = [h.median_fee for h in HOSPITALS if h.median_fee > 0]
    fallback = (
        (int(_percentile(all_medians, 25)), int(_percentile(all_medians, 75)))
        if all_medians
        else (500, 1500)
    )
    for tier in TIER_NAMES:
        TIER_FEE_BANDS.setdefault(tier, fallback)


_build_tier_model()

_HOSPITALS_BY_TIER: Dict[str, Tuple[BangaloreHospital, ...]] = {
    t: tuple(h for h in HOSPITALS if h.tier == t) for t in TIER_NAMES
}


# ---------------------------------------------------------------------------
# Symptom text -> blr.xlsx specialty department mapping
# ---------------------------------------------------------------------------
#
# The authoritative routing now comes from :mod:`specialization_mapper`, which
# analyzes the raw symptom text and returns a clinically sensible
# specialization (never defaulting to a surgical department). That
# specialization is then translated here onto the *actual* department names
# present in ``blr.xlsx`` so doctor filtering hits real rows.
#
# The dataset contains these 20 departments (see ``Hospital Directory`` sheet):
#   Cardiology, Dental, Dermatology, ENT, Endocrinology, Gastroenterology,
#   General Surgery, Gynecology, Hematology, Nephrology, Neurology, Oncology,
#   Ophthalmology, Orthopedics, Pediatrics, Plastic Surgery, Psychiatry,
#   Pulmonology, Rheumatology, Urology.
#
# Mapper specializations that have no dedicated department (General Physician,
# Internal Medicine, Physiotherapy, Nutrition/Dietetics) are routed to the
# closest clinically appropriate set of real departments so the user still gets
# relevant doctors.

# Specialty departments that actually exist in blr.xlsx.
DATASET_DEPARTMENTS: Tuple[str, ...] = (
    "Cardiology", "Dental", "Dermatology", "ENT", "Endocrinology",
    "Gastroenterology", "General Surgery", "Gynecology", "Hematology",
    "Nephrology", "Neurology", "Oncology", "Ophthalmology", "Orthopedics",
    "Pediatrics", "Plastic Surgery", "Psychiatry", "Pulmonology",
    "Rheumatology", "Urology",
)

# Map a mapper specialization onto an ordered tuple of dataset departments
# (most relevant first). Generalist/triage specializations that have no direct
# department fall back to broadly relevant outpatient departments.
_SPECIALIZATION_TO_DEPARTMENTS: Dict[str, Tuple[str, ...]] = {
    # Direct one-to-one department matches.
    "Cardiology": ("Cardiology",),
    "Dermatology": ("Dermatology",),
    "Neurology": ("Neurology",),
    "Orthopedics": ("Orthopedics", "Rheumatology"),
    "Gastroenterology": ("Gastroenterology",),
    "Endocrinology": ("Endocrinology",),
    "Pulmonology": ("Pulmonology",),
    "ENT": ("ENT",),
    "Ophthalmology": ("Ophthalmology",),
    "Gynecology": ("Gynecology",),
    "Psychiatry": ("Psychiatry",),
    "Urology": ("Urology", "Nephrology"),
    "Nephrology": ("Nephrology", "Urology"),
    "Oncology": ("Oncology", "Hematology"),
    "Pediatrics": ("Pediatrics",),
    "Rheumatology": ("Rheumatology", "Orthopedics"),
    "General Surgery": ("General Surgery",),
    "Plastic Surgery": ("Plastic Surgery",),
    "Hematology": ("Hematology", "Oncology"),
    # Dentistry maps to the dataset's "Dental" department label.
    "Dentistry": ("Dental",),
}

# Generalist / triage specializations that have NO dedicated department in
# ``blr.xlsx`` (there is no "General Physician" column, etc.). When a symptom
# maps to one of these we deliberately do NOT recommend doctors or a
# specialization — the Bangalore pipeline steps aside and the request is served
# by the generic deterministic estimator, exactly like every other Karnataka
# city (hospitals only).
GENERALIST_SPECIALIZATIONS: frozenset = frozenset(
    {
        "General Physician",
        "Internal Medicine",
        "Physiotherapy",
        "Nutrition/Dietetics",
    }
)

_MULTI_SPLIT_RE = re.compile(
    r"\s*(?:,|;|/|\band\b|\bplus\b|\balong with\b|&)+\s*",
    re.IGNORECASE,
)

# Heuristic: a Bangalore hospital with this many departments is treated as
# multi-speciality for fallback selection.
_MULTI_SPECIALTY_MIN_DEPARTMENTS = 6


def _split_symptom_phrases(text: str) -> Tuple[str, ...]:
    if not text:
        return ()
    norm = text.strip()
    if not norm:
        return ()
    parts = [p.strip(" .-") for p in _MULTI_SPLIT_RE.split(norm) if p and p.strip(" .-")]
    seen: set = set()
    cleaned: List[str] = []
    for part in parts:
        key = part.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(part)
    return tuple(cleaned)


def has_dataset_department(specialization: str) -> bool:
    """True when ``specialization`` maps to a real blr.xlsx department.

    Returns ``False`` for generalist specializations (General Physician, etc.)
    that have no dedicated department, signalling the caller to fall back to
    the generic hospitals-only pipeline.
    """
    return specialization in _SPECIALIZATION_TO_DEPARTMENTS

# Legacy fallback: condition-key -> departments (used only when the new mapper
# is unavailable for some reason). Note "general" deliberately routes to
# outpatient departments, NOT General Surgery.
_LEGACY_CONDITION_TO_SPECIALTIES: Dict[str, Tuple[str, ...]] = {
    "general": ("Pediatrics", "Dermatology", "ENT", "Gastroenterology"),
    "diabetes": ("Endocrinology",),
    "hypertension": ("Cardiology", "Nephrology"),
    "asthma": ("Pulmonology", "ENT"),
    "dental": ("Dental",),
    "ophthalmology": ("Ophthalmology",),
    "orthopedics": ("Orthopedics", "Rheumatology"),
    "cardiology": ("Cardiology",),
    "neurology": ("Neurology",),
    "gynaecology": ("Gynecology",),
    "oncology": ("Oncology", "Hematology"),
}


def map_symptom_to_specialization(symptom_text: str) -> str:
    """Return the best mapper specialization for a free-text complaint.

    Thin wrapper over :mod:`specialization_mapper` with a safe default of
    ``"General Physician"`` (never a surgical department).
    """
    if _MAPPER_AVAILABLE and _spec_mapper is not None:
        try:
            return _spec_mapper.best_specialization(symptom_text)
        except Exception:  # pragma: no cover - defensive
            logger.exception("bangalore_estimator: specialization_mapper failed")
    return "General Physician"


def map_symptom_to_specializations(symptom_text: str) -> Tuple[str, ...]:
    """Return unique mapper specializations for multi-symptom text."""
    phrases = _split_symptom_phrases(symptom_text) or (symptom_text,)
    seen: set = set()
    ordered: List[str] = []
    for phrase in phrases:
        spec = map_symptom_to_specialization(phrase)
        key = spec.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(spec)
    return tuple(ordered)


def specialties_for_text(symptom_text: str) -> Tuple[str, ...]:
    """Map raw symptom text to an ordered tuple of blr.xlsx departments.

    Returns an empty tuple when the symptom maps to a generalist
    specialization that has no dedicated department in the dataset (the caller
    then serves hospitals only, like other Karnataka cities).
    """
    specs = map_symptom_to_specializations(symptom_text)
    ordered: List[str] = []
    seen: set = set()
    for spec in specs:
        for dept in _SPECIALIZATION_TO_DEPARTMENTS.get(spec, ()):
            if dept in seen:
                continue
            seen.add(dept)
            ordered.append(dept)
    return tuple(ordered)


def condition_specialties(condition_key: str) -> Tuple[str, ...]:
    """Legacy condition-key -> department mapping (kept for compatibility).

    Prefer :func:`specialties_for_text` for new code. This remains for any
    caller that only has a resolved condition key. It never returns
    ``General Surgery`` for the generic ``"general"`` key.
    """
    return _LEGACY_CONDITION_TO_SPECIALTIES.get(condition_key, ())


# ---------------------------------------------------------------------------
# Number of consultations by severity (drives the fee-anchored consult line)
# ---------------------------------------------------------------------------

_VISITS: Dict[str, Tuple[int, int]] = {
    "mild": (1, 2),
    "moderate": (2, 3),
    "severe": (3, 5),
}


# ---------------------------------------------------------------------------
# Doctor / hospital relevance + selection
# ---------------------------------------------------------------------------


def _doctor_relevance(doc: Doctor, target_specialties: Tuple[str, ...]) -> float:
    if not target_specialties:
        base = 0.45
    elif doc.specialty in target_specialties:
        # Primary specialty match scores highest; later entries slightly less.
        rank = target_specialties.index(doc.specialty)
        base = 1.0 - (rank * 0.12)
    else:
        base = 0.2
    # Experience nudges relevance up a touch (capped).
    base += min(0.15, doc.experience_years / 200.0)
    return round(min(1.0, base), 3)


def _hospital_relevance(
    hospital: BangaloreHospital, target_specialties: Tuple[str, ...]
) -> float:
    has_specialty = any(s in hospital.specialties for s in target_specialties)
    if target_specialties and has_specialty:
        score = 0.8
    elif not target_specialties:
        score = 0.5
    else:
        score = 0.3
    score += max(0.0, min(0.2, (hospital.rating - 3.5) * 0.12))
    return round(min(1.0, score), 3)


def _select_doctors(
    hospital: BangaloreHospital,
    target_specialties: Tuple[str, ...],
    limit: int = 3,
) -> List[Tuple[Doctor, float]]:
    scored = [
        (d, _doctor_relevance(d, target_specialties)) for d in hospital.doctors
    ]
    # Prefer relevant specialty, then experience, then lower fee (accessible).
    scored.sort(
        key=lambda pair: (
            -pair[1],
            -pair[0].experience_years,
            pair[0].consultation_fee,
        )
    )
    # Keep only doctors with some relevance when we have a target; otherwise
    # fall back to the top few by experience.
    relevant = [pair for pair in scored if pair[1] >= 0.5]
    chosen = relevant if relevant else scored
    return chosen[:limit]


def _is_multi_speciality(hospital: BangaloreHospital) -> bool:
    return len(hospital.specialties) >= _MULTI_SPECIALTY_MIN_DEPARTMENTS


def _select_doctors_multi(
    hospital: BangaloreHospital,
    target_specialties: Tuple[str, ...],
    limit: int,
) -> List[Tuple[Doctor, float]]:
    """Pick at least one doctor per specialty when possible."""
    chosen: List[Tuple[Doctor, float]] = []
    seen: set = set()

    for spec in target_specialties:
        docs = [d for d in hospital.doctors if d.specialty == spec]
        if not docs:
            continue
        docs.sort(key=lambda d: (-d.experience_years, d.consultation_fee))
        doc = docs[0]
        if doc.name in seen:
            continue
        chosen.append((doc, _doctor_relevance(doc, target_specialties)))
        seen.add(doc.name)

    if len(chosen) < limit:
        extras = _select_doctors(hospital, target_specialties, limit=limit)
        for doc, rel in extras:
            if doc.name in seen:
                continue
            chosen.append((doc, rel))
            seen.add(doc.name)
            if len(chosen) >= limit:
                break
    return chosen

def _stratified_pick(
    candidates: List[Tuple[BangaloreHospital, float]], *, target: int
) -> List[Tuple[BangaloreHospital, float]]:
    """Distribute the pick across Low/Mid/High tiers for a balanced spread."""
    if not candidates:
        return []
    buckets: Dict[str, List[Tuple[BangaloreHospital, float]]] = {t: [] for t in TIER_NAMES}
    for h, rel in candidates:
        buckets.setdefault(h.tier, []).append((h, rel))
    present = [t for t in TIER_NAMES if buckets[t]]
    picked: List[Tuple[BangaloreHospital, float]] = []
    while len(picked) < target and present:
        for tier in list(present):
            if not buckets[tier]:
                present.remove(tier)
                continue
            picked.append(buckets[tier].pop(0))
            if len(picked) >= target:
                break
    return picked


def match_hospitals(
    *,
    target_specialties: Tuple[str, ...],
    requested_tier: Optional[str],
    limit: int = 6,
) -> List[Tuple[BangaloreHospital, float]]:
    """Rank Bangalore hospitals by specialization relevance + rating."""
    pool: Tuple[BangaloreHospital, ...]
    if requested_tier:
        pool = _HOSPITALS_BY_TIER.get(requested_tier, ())
        if not pool:
            pool = HOSPITALS
    else:
        pool = HOSPITALS

    scored: List[Tuple[BangaloreHospital, float]] = []
    for h in pool:
        rel = _hospital_relevance(h, target_specialties)
        scored.append((h, rel))

    scored.sort(key=lambda pair: (-pair[1], -pair[0].rating, pair[0].name))

    if requested_tier:
        return scored[:limit]
    return _stratified_pick(scored, target=limit)


def match_hospitals_multi(
    *,
    target_specialties: Tuple[str, ...],
    requested_tier: Optional[str],
    limit: int = 6,
) -> Tuple[List[Tuple[BangaloreHospital, float]], str, Dict[int, Tuple[str, ...]]]:
    """Match hospitals for multiple specialties, preferring single-hospital coverage."""
    pool: Tuple[BangaloreHospital, ...]
    if requested_tier:
        pool = _HOSPITALS_BY_TIER.get(requested_tier, ())
        if not pool:
            pool = HOSPITALS
    else:
        pool = HOSPITALS

    if not pool or not target_specialties:
        return [], "none", {}

    full_matches: List[Tuple[BangaloreHospital, float]] = []
    for h in pool:
        if all(s in h.specialties for s in target_specialties):
            full_matches.append((h, _hospital_relevance(h, target_specialties)))

    if full_matches:
        full_matches.sort(key=lambda pair: (-pair[1], -pair[0].rating, pair[0].name))
        picked = _stratified_pick(full_matches, target=limit)
        focus = {h.hospital_id: target_specialties for h, _ in picked}
        return picked, "single_hospital", focus

    by_spec: Dict[str, List[Tuple[BangaloreHospital, float]]] = {}
    missing_specs: List[str] = []
    for spec in target_specialties:
        matches = [
            (h, _hospital_relevance(h, (spec,)))
            for h in pool
            if spec in h.specialties
        ]
        if not matches:
            missing_specs.append(spec)
            continue
        matches.sort(key=lambda pair: (-pair[1], -pair[0].rating, pair[0].name))
        by_spec[spec] = matches

    if missing_specs:
        multi_pool = [(h, _hospital_relevance(h, target_specialties)) for h in pool if _is_multi_speciality(h)]
        if not multi_pool:
            multi_pool = [(h, _hospital_relevance(h, target_specialties)) for h in pool]
        multi_pool.sort(key=lambda pair: (-pair[1], -pair[0].rating, pair[0].name))
        picked = _stratified_pick(multi_pool, target=limit)
        focus = {h.hospital_id: target_specialties for h, _ in picked}
        return picked, "multi_speciality", focus

    combined: List[Tuple[BangaloreHospital, float]] = []
    focus: Dict[int, Tuple[str, ...]] = {}
    seen: set = set()
    for spec in target_specialties:
        for hospital, rel in by_spec.get(spec, []):
            if hospital.hospital_id in seen:
                continue
            combined.append((hospital, rel))
            focus[hospital.hospital_id] = (spec,)
            seen.add(hospital.hospital_id)
            break

    if len(combined) < limit:
        remaining: List[Tuple[BangaloreHospital, float]] = []
        for spec in target_specialties:
            remaining.extend(by_spec.get(spec, [])[1:])
        remaining.sort(key=lambda pair: (-pair[1], -pair[0].rating, pair[0].name))
        for hospital, rel in remaining:
            if hospital.hospital_id in seen:
                continue
            combined.append((hospital, rel))
            focus[hospital.hospital_id] = (hospital.specialties[0],) if hospital.specialties else ()
            seen.add(hospital.hospital_id)
            if len(combined) >= limit:
                break

    return combined[:limit], "multiple_hospitals", focus


# ---------------------------------------------------------------------------
# Cost model (fee-anchored)
# ---------------------------------------------------------------------------

# Map Bangalore tier label -> the internal tier used by cost_estimator's
# calibrated structural table (which uses Low/Medium/High).
_TIER_TO_INTERNAL = {"Low": "Low", "Mid": "Medium", "High": "High"}

_ENVELOPE_LOW_MULT = 0.65
_ENVELOPE_HIGH_MULT = 1.40


def _consultation_band(
    *, tier: str, severity: str, consultation_mult: float, fee_override: Optional[Tuple[int, int]] = None
) -> Tuple[float, float]:
    """Fee-anchored consultation cost band for a tier.

    The band is the real dataset fee range for the tier, multiplied by the
    plausible number of visits for the severity and the consultation-type
    multiplier.
    """
    fee_lo, fee_hi = fee_override if fee_override else TIER_FEE_BANDS.get(tier, (500, 1500))
    v_lo, v_hi = _VISITS.get(severity, (2, 3))
    lo = fee_lo * v_lo * consultation_mult
    hi = fee_hi * v_hi * consultation_mult
    if hi <= lo:
        hi = lo * 1.2 + 1
    return lo, hi


def _structural_other_lines(
    *, condition_key: str, internal_tier: str, severity: str
) -> Dict[str, Tuple[float, float]]:
    """Tests/medication/procedure/hospitalization bands from the calibrated
    structural model (everything except consultation)."""
    total_lo, total_hi = cost_estimator._resolve_total_range(
        condition_key, internal_tier, severity
    )
    weights = cost_estimator._resolve_weights(condition_key, severity)
    wtot = float(sum(weights.values())) or 1.0
    out: Dict[str, Tuple[float, float]] = {}
    for line, w in weights.items():
        if line == "consultation":
            continue
        share = w / wtot
        out[line] = (total_lo * share, total_hi * share)
    return out


def _compose_breakdown(
    *,
    condition_key: str,
    tier: str,
    severity: str,
    consultation_mult: float,
    fee_override: Optional[Tuple[int, int]] = None,
) -> Tuple[Dict[str, Tuple[float, float]], float, float]:
    """Build the full line breakdown (raw, unrounded) and total for a tier."""
    internal_tier = _TIER_TO_INTERNAL.get(tier, "Medium")
    breakdown = _structural_other_lines(
        condition_key=condition_key, internal_tier=internal_tier, severity=severity
    )
    breakdown["consultation"] = _consultation_band(
        tier=tier,
        severity=severity,
        consultation_mult=consultation_mult,
        fee_override=fee_override,
    )
    total_lo = sum(b[0] for b in breakdown.values())
    total_hi = sum(b[1] for b in breakdown.values())
    return breakdown, total_lo, total_hi


def _estimate_condition_breakdown(
    *,
    condition_key: str,
    tier: Optional[str],
    severity: str,
    consultation_mult: float,
    present_tiers: Tuple[str, ...],
    auto_tier: bool,
) -> Tuple[Dict[str, Tuple[float, float]], float, float, Dict[str, Tuple[float, float]]]:
    """Return raw breakdown + totals for one condition."""
    tier_breakdown: Dict[str, Tuple[float, float]] = {}
    if auto_tier:
        ordered = [t for t in TIER_NAMES if t in present_tiers] or list(TIER_NAMES)
        low_bd, low_total_lo, _ = _compose_breakdown(
            condition_key=condition_key,
            tier=ordered[0],
            severity=severity,
            consultation_mult=consultation_mult,
        )
        high_bd, _, high_total_hi = _compose_breakdown(
            condition_key=condition_key,
            tier=ordered[-1],
            severity=severity,
            consultation_mult=consultation_mult,
        )
        breakdown_raw = {
            line: (low_bd.get(line, (0.0, 0.0))[0], high_bd.get(line, (0.0, 0.0))[1])
            for line in set(low_bd) | set(high_bd)
        }
        total_min = float(low_total_lo)
        total_max = float(high_total_hi)
        for t in present_tiers:
            _, t_lo, t_hi = _compose_breakdown(
                condition_key=condition_key,
                tier=t,
                severity=severity,
                consultation_mult=consultation_mult,
            )
            tier_breakdown[t] = (t_lo, t_hi)
    else:
        assert tier is not None
        breakdown_raw, total_min, total_max = _compose_breakdown(
            condition_key=condition_key,
            tier=tier,
            severity=severity,
            consultation_mult=consultation_mult,
        )
    return breakdown_raw, total_min, total_max, tier_breakdown


def _hospital_estimated_range(
    *,
    hospital: BangaloreHospital,
    condition_key: str,
    severity: str,
    consultation_mult: float,
    selected_doctors: List[Tuple[Doctor, float]],
) -> Tuple[int, int]:
    """Per-hospital estimated total treatment range, anchored on that
    hospital's own consultation fees."""
    if selected_doctors:
        fees = [d.consultation_fee for d, _ in selected_doctors if d.consultation_fee > 0]
    else:
        fees = [d.consultation_fee for d in hospital.doctors if d.consultation_fee > 0]
    if fees:
        fee_override = (min(fees), max(fees))
    else:
        fee_override = TIER_FEE_BANDS.get(hospital.tier, (500, 1500))

    _, lo, hi = _compose_breakdown(
        condition_key=condition_key,
        tier=hospital.tier,
        severity=severity,
        consultation_mult=consultation_mult,
        fee_override=fee_override,
    )
    rlo, rhi = cost_estimator._round_band(lo, hi)
    return rlo, rhi


def _hospital_estimated_range_multi(
    *,
    hospital: BangaloreHospital,
    conditions: Tuple[ConditionProfile, ...],
    severity: str,
    consultation_mult: float,
    selected_doctors: List[Tuple[Doctor, float]],
) -> Tuple[int, int]:
    if selected_doctors:
        fees = [d.consultation_fee for d, _ in selected_doctors if d.consultation_fee > 0]
    else:
        fees = [d.consultation_fee for d in hospital.doctors if d.consultation_fee > 0]
    if fees:
        fee_override = (min(fees), max(fees))
    else:
        fee_override = TIER_FEE_BANDS.get(hospital.tier, (500, 1500))

    total_lo = 0.0
    total_hi = 0.0
    for cond in conditions:
        _, lo, hi = _compose_breakdown(
            condition_key=cond.key,
            tier=hospital.tier,
            severity=severity,
            consultation_mult=consultation_mult,
            fee_override=fee_override,
        )
        total_lo += lo
        total_hi += hi
    rlo, rhi = cost_estimator._round_band(total_lo, total_hi)
    return rlo, rhi


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def estimate(
    *,
    city: str,
    condition_text: str,
    severity: Optional[str] = None,
    hospital_tier: Optional[str] = None,
    consultation_type: Optional[str] = None,
) -> Dict[str, object]:
    """Generate a Bangalore-specific, fee-driven cost estimate.

    The return shape is a superset of :func:`cost_estimator.estimate` so the
    API layer can compose the response uniformly. Additional Bangalore-only
    keys (``bangalore_mode``, doctor recommendations, per-hospital fee and
    cost ranges, tier classification) make the response richer than the
    generic pipeline.
    """
    conditions: Tuple[ConditionProfile, ...] = cost_estimator.resolve_conditions(condition_text)
    primary_condition = conditions[0]

    severity_norm = (severity or "moderate").strip().lower()
    if severity_norm not in ("mild", "moderate", "severe"):
        severity_norm = "moderate"

    consultation_norm = (
        (consultation_type or "specialist").strip().lower().replace("-", "_").replace(" ", "_")
    )
    if consultation_norm not in cost_estimator.CONSULTATION_MULT:
        consultation_norm = "specialist"
    consult_mult = cost_estimator.CONSULTATION_MULT.get(consultation_norm, 1.0)

    # Normalize requested tier onto the Bangalore Low/Mid/High vocabulary.
    requested_tier: Optional[str] = None
    if hospital_tier and str(hospital_tier).strip():
        ht = str(hospital_tier).strip().lower()
        if ht.startswith("low"):
            requested_tier = "Low"
        elif ht.startswith("hi"):
            requested_tier = "High"
        elif ht.startswith("med") or ht == "mid":
            requested_tier = "Mid"
    auto_tier = requested_tier is None

    # Route the raw symptom text through the production specialization mapper
    # (fixes the legacy "fever -> General Surgery" misrouting). The chosen
    # specialization is surfaced in the response and translated onto real
    # blr.xlsx departments for doctor/hospital matching.
    mapped_specializations = map_symptom_to_specializations(condition_text)
    mapped_specialization = (
        mapped_specializations[0] if len(mapped_specializations) == 1 else None
    )

    target_specialties = specialties_for_text(condition_text)

    # If there are no dataset departments to match (generalist-only symptoms),
    # fall back to the generic deterministic estimator.
    if not target_specialties:
        return cost_estimator.estimate(
            city=city,
            condition_text=condition_text,
            severity=severity,
            hospital_tier=hospital_tier,
            consultation_type=consultation_type,
        )

    # ---- Hospital + doctor recommendations -------------------------------
    if len(target_specialties) > 1:
        matches, match_strategy, focus_map = match_hospitals_multi(
            target_specialties=target_specialties,
            requested_tier=requested_tier,
            limit=6,
        )
    else:
        matches = match_hospitals(
            target_specialties=target_specialties,
            requested_tier=requested_tier,
            limit=6,
        )
        match_strategy = "single"
        focus_map = {h.hospital_id: target_specialties for h, _ in matches}

    # ---- Present tiers ---------------------------------------------------
    if auto_tier:
        present_tiers = tuple(t for t in TIER_NAMES if _HOSPITALS_BY_TIER.get(t))
        if not present_tiers:
            present_tiers = TIER_NAMES
    else:
        present_tiers = (requested_tier,)  # type: ignore[assignment]

    breakdown_raw: Dict[str, Tuple[float, float]] = {}
    tier_breakdown_raw: Dict[str, Tuple[float, float]] = {
        t: (0.0, 0.0) for t in present_tiers
    } if auto_tier else {}

    for cond in conditions:
        cond_breakdown, _, _, cond_tiers = _estimate_condition_breakdown(
            condition_key=cond.key,
            tier=requested_tier,
            severity=severity_norm,
            consultation_mult=consult_mult,
            present_tiers=present_tiers,
            auto_tier=auto_tier,
        )
        for line, (lo, hi) in cond_breakdown.items():
            base = breakdown_raw.get(line, (0.0, 0.0))
            breakdown_raw[line] = (base[0] + lo, base[1] + hi)
        if auto_tier:
            for tier, (lo, hi) in cond_tiers.items():
                base = tier_breakdown_raw.get(tier, (0.0, 0.0))
                tier_breakdown_raw[tier] = (base[0] + lo, base[1] + hi)

    # ---- Rounding --------------------------------------------------------
    breakdown_rounded: Dict[str, Dict[str, int]] = {}
    for line, (lo, hi) in breakdown_raw.items():
        rlo, rhi = cost_estimator._round_band(lo, hi)
        breakdown_rounded[line] = {"min": rlo, "max": rhi}
    total_min = sum(lo for lo, _ in breakdown_raw.values())
    total_max = sum(hi for _, hi in breakdown_raw.values())
    rounded_total_min, rounded_total_max = cost_estimator._round_band(total_min, total_max)

    # ---- Per-tier breakdown ---------------------------------------------
    tier_breakdown: Dict[str, Dict[str, int]] = {}
    if auto_tier:
        for tier, (lo, hi) in tier_breakdown_raw.items():
            r_lo, r_hi = cost_estimator._round_band(lo, hi)
            tier_breakdown[tier] = {"min": r_lo, "max": r_hi}

    # ---- Allowed envelope (hard constraint for Gemini refinement) -------
    allowed_total_min, allowed_total_max = cost_estimator._round_band(
        max(0.0, total_min * _ENVELOPE_LOW_MULT), total_max * _ENVELOPE_HIGH_MULT
    )
    allowed_components: Dict[str, Dict[str, int]] = {}
    for line, (lo, hi) in breakdown_raw.items():
        a_lo, a_hi = cost_estimator._round_band(
            max(0.0, lo * _ENVELOPE_LOW_MULT), hi * _ENVELOPE_HIGH_MULT
        )
        allowed_components[line] = {"min": a_lo, "max": a_hi}

    # ---- Hospital payload with doctor recommendations -------------------
    matched_hospitals_payload: List[Dict[str, object]] = []
    for hospital, rel in matches:
        focus_specialties = focus_map.get(hospital.hospital_id, target_specialties)
        if len(focus_specialties) > 1 and match_strategy == "single_hospital":
            selected = _select_doctors_multi(
                hospital,
                focus_specialties,
                limit=max(3, len(focus_specialties)),
            )
        else:
            selected = _select_doctors(hospital, focus_specialties, limit=3)

        if len(conditions) > 1:
            est_lo, est_hi = _hospital_estimated_range_multi(
                hospital=hospital,
                conditions=conditions,
                severity=severity_norm,
                consultation_mult=consult_mult,
                selected_doctors=selected,
            )
        else:
            est_lo, est_hi = _hospital_estimated_range(
                hospital=hospital,
                condition_key=primary_condition.key,
                severity=severity_norm,
                consultation_mult=consult_mult,
                selected_doctors=selected,
            )
        doctor_payload = [
            {
                "name": d.name,
                "specialization": d.specialty,
                "qualification": d.qualification or None,
                "experience_years": d.experience_years or None,
                "consultation_fee": d.consultation_fee or None,
                "availability": d.availability or None,
                "timing": d.timing or None,
            }
            for d, _ in selected
        ]
        # Consultation fee range across the recommended doctors.
        sel_fees = [d.consultation_fee for d, _ in selected if d.consultation_fee > 0]
        if len(focus_specialties) > 1:
            primary_specialization = "Multi-speciality"
        else:
            primary_specialization = (
                selected[0][0].specialty
                if selected
                else (hospital.specialties[0] if hospital.specialties else "Multi-speciality")
            )
        matched_hospitals_payload.append(
            {
                "name": hospital.name,
                "city": "Bengaluru",
                "district": hospital.area or "Bengaluru Urban",
                "hospital_type": hospital.hospital_type,
                "specialization": primary_specialization,
                "rating": round(hospital.rating, 1),
                "cost_level": hospital.tier,
                "relevance_score": rel,
                # Bangalore-specific enrichment:
                "area": hospital.area or None,
                "tier": hospital.tier,
                "accreditation": hospital.accreditation or None,
                "total_beds": hospital.total_beds or None,
                "consultation_fee_min": (min(sel_fees) if sel_fees else hospital.min_fee) or None,
                "consultation_fee_max": (max(sel_fees) if sel_fees else hospital.max_fee) or None,
                "estimated_cost_min": est_lo,
                "estimated_cost_max": est_hi,
                "doctors": doctor_payload,
            }
        )

    # ---- Narrative -------------------------------------------------------
    confidence_note = _build_confidence_note(
        matches=matches,
        condition=primary_condition,
        auto=auto_tier,
        present_tiers=present_tiers,
        requested_tier=requested_tier,
    )
    relevance_summary = _build_relevance_summary(
        matches=matches,
        condition=primary_condition,
        target_specialties=target_specialties,
        match_strategy=match_strategy,
        condition_count=len(conditions),
    )

    display_tier = "Auto" if auto_tier else requested_tier

    return {
        "city": "Bengaluru",
        "condition": (
            {"key": primary_condition.key, "label": primary_condition.label}
            if len(conditions) == 1
            else {"key": "multi", "label": "Multiple conditions"}
        ),
        "conditions": [{"key": c.key, "label": c.label} for c in conditions],
        "tier": display_tier,
        "severity": severity_norm,
        "consultation_type": consultation_norm,
        "estimated_total_min": rounded_total_min,
        "estimated_total_max": rounded_total_max,
        "breakdown": breakdown_rounded,
        "tier_breakdown": tier_breakdown if auto_tier else {},
        "present_tiers": list(present_tiers),
        "allowed_range": {"min": allowed_total_min, "max": allowed_total_max},
        "allowed_components": allowed_components,
        "matched_hospitals": matched_hospitals_payload,
        "confidence_note": confidence_note,
        "relevance_summary": relevance_summary,
        "bangalore_mode": True,
        "mapped_specialization": mapped_specialization,
        "mapped_specializations": list(mapped_specializations),
        "target_departments": list(target_specialties),
        "tier_thresholds": {
            "low_max_fee": TIER_FEE_BANDS.get("Low", (0, 0))[1],
            "high_min_fee": TIER_FEE_BANDS.get("High", (0, 0))[0],
        },
    }


# ---------------------------------------------------------------------------
# Narrative helpers
# ---------------------------------------------------------------------------


def _build_confidence_note(
    *,
    matches: List[Tuple[BangaloreHospital, float]],
    condition: ConditionProfile,
    auto: bool,
    present_tiers: Tuple[str, ...],
    requested_tier: Optional[str],
) -> str:
    doctor_count = sum(len(h.doctors) for h, _ in matches)
    if not matches:
        return (
            "No Bangalore hospitals matched this condition in the directory; "
            "the range uses fee-driven base bands as a reference."
        )
    if auto:
        tier_list = ", ".join(t.lower() for t in present_tiers)
        return (
            f"Localized Bangalore estimate spanning the {tier_list} pricing tiers, "
            f"informed by {len(matches)} hospitals and {doctor_count} doctors with "
            f"real consultation fees. Pick a tier above to narrow it."
        )
    tier_label = (requested_tier or "Mid").lower()
    return (
        f"Estimated against {tier_label}-tier Bangalore hospitals, anchored on "
        f"actual consultation fees from {len(matches)} hospitals and {doctor_count} "
        f"doctors. Actual costs vary by doctor and treatment plan."
    )


def _build_relevance_summary(
    *,
    matches: List[Tuple[BangaloreHospital, float]],
    condition: ConditionProfile,
    target_specialties: Tuple[str, ...],
    match_strategy: str = "single",
    condition_count: int = 1,
) -> str:
    if not matches:
        if condition_count > 1:
            return "No Bangalore hospitals matched these conditions yet."
        return f"No {condition.label.lower()} specialists indexed for Bengaluru yet."

    parts = [f"{len(matches)} Bangalore hospitals reviewed"]

    if condition_count > 1:
        if target_specialties:
            parts.append(f"{len(target_specialties)} specialties requested")
        if match_strategy == "multiple_hospitals":
            parts.append("different hospitals per specialty")
        elif match_strategy == "multi_speciality":
            parts.append("multi-speciality focus")
        elif match_strategy == "single_hospital":
            parts.append("single hospital covers all specialties")
        return "; ".join(parts) + "."

    spec_label = target_specialties[0] if target_specialties else condition.label
    spec_hits = sum(
        1 for h, _ in matches if any(s in h.specialties for s in target_specialties)
    )
    if target_specialties and spec_hits:
        parts.append(f"{spec_hits} with a {spec_label.lower()} department")
    return "; ".join(parts) + "."


__all__ = [
    "AVAILABLE",
    "HOSPITALS",
    "TIER_NAMES",
    "TIER_FEE_BANDS",
    "Doctor",
    "BangaloreHospital",
    "is_bangalore",
    "condition_specialties",
    "specialties_for_text",
    "map_symptom_to_specialization",
    "has_dataset_department",
    "GENERALIST_SPECIALIZATIONS",
    "DATASET_DEPARTMENTS",
    "match_hospitals",
    "estimate",
]
